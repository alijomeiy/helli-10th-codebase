"""activities.py — grading hub blueprint.

Aggregates class activities (auto-synced CTF results + manually scored
quizzes) per student, scoped by teacher. Provides:
  - /my                      student's own report card (charts + table)
  - /admin/hub               hub home (activities list + class averages)
  - /admin/hub/quiz/*        create / grid-score / xlsx template / upload / delete
  - /admin/hub/ctf/sync      pull per-student solves from the CTFd API
  - /admin/hub/export        whole-school (or per-teacher) xlsx report
  - /admin/hub/teachers      super-admin: teacher accounts + assignments
"""
import io
import json
import re
from datetime import datetime, date
from functools import wraps

import requests
from flask import (
    Blueprint, render_template, redirect, url_for, request, flash, abort,
    send_file, current_app,
)
from flask_login import login_required, current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from models import db, User, Activity, Score, STATUS_APPROVED

hub_bp = Blueprint("hub", __name__, url_prefix="/admin/hub")
my_bp = Blueprint("my", __name__)

CTF_ACTIVITY_TITLE = "مسابقه پرچم‌یابی (CTF)"


# ---------- helpers ------------------------------------------------------------

def is_super():
    return bool(current_user.is_authenticated and current_user.role == "admin"
                and current_user.is_super)


def hub_required(f):
    """Any admin (teacher or super) may use the hub."""
    @wraps(f)
    def wrap(*a, **kw):
        if not current_user.is_authenticated:
            return redirect(url_for("admin.admin_login"))
        if current_user.role != "admin":
            abort(403)
        return f(*a, **kw)
    return wrap


def super_required(f):
    @wraps(f)
    def wrap(*a, **kw):
        if not current_user.is_authenticated:
            return redirect(url_for("admin.admin_login"))
        if current_user.role != "admin" or not current_user.is_super:
            abort(403)
        return f(*a, **kw)
    return wrap


def scoped_students():
    """Students visible to the current admin (teacher scope or all for super)."""
    q = User.query.filter(User.role == "student", User.status == STATUS_APPROVED)
    if not current_user.is_super:
        q = q.filter(User.teacher_id == current_user.id)
    return q.order_by(User.username).all()


def teacher_name(teacher_id, teachers):
    t = teachers.get(teacher_id)
    return t.full_name or t.username if t else "—"


def activity_rows(students, activities):
    """Matrix data: {(activity_id, user_id): Score} for the given students."""
    if not students or not activities:
        return {}
    uids = [s.id for s in students]
    aids = [a.id for a in activities]
    rows = Score.query.filter(Score.user_id.in_(uids),
                              Score.activity_id.in_(aids)).all()
    return {(r.activity_id, r.user_id): r for r in rows}


def fmt_score(sc, max_score):
    if sc is None:
        return "—"
    s = int(sc) if float(sc).is_integer() else sc
    m = int(max_score) if float(max_score).is_integer() else max_score
    return f"{s} / {m}"


# ---------- CTFd sync ----------------------------------------------------------

def ctfd_get_all(cfg, path, params=None):
    """GET every page of a CTFd API resource. Token auth needs the JSON
    content-type header even for GETs (CTFd quirk)."""
    headers = {"Authorization": f"Token {cfg['token']}",
               "Content-Type": "application/json"}
    out, page = [], 1
    while True:
        p = dict(params or {})
        p["page"] = page
        r = requests.get(f"{cfg['url']}/api/v1{path}",
                         headers=headers, params=p, timeout=30)
        r.raise_for_status()
        data = r.json().get("data", [])
        out.extend(data)
        nxt = (r.json().get("meta", {}).get("pagination", {}) or {}).get("next")
        if not nxt or not data:
            break
        page += 1
    return out


def ctfd_config():
    url = (current_app.config.get("CTFD_URL") or "").rstrip("/")
    token = current_app.config.get("CTFD_TOKEN") or ""
    if not url or not token:
        return None
    return {"url": url, "token": token}


def sync_ctf_activity():
    """Pull all correct submissions from CTFd and upsert the ctf Activity."""
    cfg = ctfd_config()
    if not cfg:
        flash("تنظیمات CTFd (CTFD_URL / CTFD_TOKEN) کامل نیست.", "error")
        return redirect(url_for("hub.hub_home"))

    challenges = {c["id"]: c for c in ctfd_get_all(cfg, "/challenges")}
    users_raw = ctfd_get_all(cfg, "/users", {"view": "admin"})
    ctfd_users = {u["id"]: u["name"] for u in users_raw}
    by_name = {v: k for k, v in ctfd_users.items()}

    # aggregate correct submissions per ctfd user
    solves = {}   # ctfd_uid -> {challenge_id: iso_date}
    for s in ctfd_get_all(cfg, "/submissions", {"type": "correct"}):
        uid = s.get("user_id")
        cid = s.get("challenge_id")
        if uid is None or cid is None:
            nested = s.get("user") or {}
            uid = uid if uid is not None else nested.get("id")
            nested_c = s.get("challenge") or {}
            cid = cid if cid is not None else nested_c.get("id")
        if uid is None or cid is None:
            continue
        solves.setdefault(uid, {})[cid] = s.get("date")

    total_possible = sum(float(c.get("value") or 0) for c in challenges.values())

    act = Activity.query.filter_by(kind="ctf").first()
    if not act:
        act = Activity(kind="ctf", title=CTF_ACTIVITY_TITLE,
                       held_on=date.today(), max_score=total_possible or 100,
                       created_by=current_user.id)
        db.session.add(act)
    else:
        act.max_score = total_possible or act.max_score
    db.session.flush()

    # map panel students -> ctfd user via username
    students = User.query.filter(User.role == "student",
                                 User.status == STATUS_APPROVED).all()
    updated = 0
    for st in students:
        cu = by_name.get(st.username)
        if cu is None:
            continue
        got = solves.get(cu, {})
        detail = sorted(
            [{"name": challenges[cid]["name"],
              "points": float(challenges[cid].get("value") or 0),
              "date": d}
             for cid, d in got.items() if cid in challenges],
            key=lambda x: x["date"] or "")
        row = Score.query.filter_by(activity_id=act.id, user_id=st.id).first()
        if not row:
            row = Score(activity_id=act.id, user_id=st.id)
            db.session.add(row)
        row.score = sum(d["points"] for d in detail)
        row.detail = json.dumps(detail, ensure_ascii=False)
        row.updated_by = current_user.id
        updated += 1
    db.session.commit()
    flash(f"نتایج CTF همگام شد: {updated} دانشجو، {len(challenges)} چالش، "
          f"سقف امتیاز {int(total_possible)}.", "ok")
    return redirect(url_for("hub.hub_home"))


# ---------- Excel ---------------------------------------------------------------

HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF")


def _sheet_header(ws, headers):
    ws.append(headers)
    for c in ws[1]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _persian_ws_title(wb, title):
    """openpyxl sheet titles forbid some chars and >31 length; keep it safe."""
    t = re.sub(r"[\\/*?:\[\]]", " ", title)[:31] or "Sheet"
    if t.lower() in (ws.title.lower() for ws in wb.worksheets):
        t = f"{t[:28]} 2"
    return wb.create_sheet(t)


def parse_score_cell(v):
    """Accept 3, 3.5, '3/5', '3 از 5' -> float or None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+(?:\.\d+)?", str(v).replace("٫", "."))
    return float(m.group()) if m else None


def build_workbook(students, activities, scores, teachers):
    """Whole report: summary matrix + one detail sheet per activity."""
    wb = Workbook()
    ws = wb.active
    ws.title = "کارنامه"
    headers = ["شماره دانش‌آموزی", "نام کاربری", "نام و نام خانوادگی", "معلم"]
    for a in activities:
        headers.append(f"{a.title} (از {fmt_int(a.max_score)})")
    headers.append("مجموع")
    _sheet_header(ws, headers)

    for st in students:
        row = [st.student_no or "—", st.username, st.full_name,
               teacher_name(st.teacher_id, teachers)]
        total = 0.0
        for a in activities:
            sc = scores.get((a.id, st.id))
            row.append(sc.score if sc else 0)
            total += sc.score if sc else 0
        row.append(round(total, 2))
        ws.append(row)

    # detail sheet per activity
    for a in activities:
        d = _persian_ws_title(wb, f"جزئیات {a.title}")
        if a.kind == "ctf":
            names = []
            for st in students:
                sc = scores.get((a.id, st.id))
                if sc and sc.detail:
                    for it in json.loads(sc.detail):
                        if it["name"] not in names:
                            names.append(it["name"])
            _sheet_header(d, ["نام کاربری", "نام", *names, "مجموع"])
            for st in students:
                sc = scores.get((a.id, st.id))
                det = {it["name"]: it["points"] for it in
                       (json.loads(sc.detail) if sc and sc.detail else [])}
                d.append([st.username, st.full_name,
                          *[det.get(n, "—") for n in names],
                          sc.score if sc else 0])
        else:
            _sheet_header(d, ["نام کاربری", "نام", "نمره", "توضیح"])
            for st in students:
                sc = scores.get((a.id, st.id))
                note = (sc.detail or "") if sc else ""
                d.append([st.username, st.full_name,
                          sc.score if sc else "—", note])
    return wb


def fmt_int(x):
    return int(x) if float(x).is_integer() else x


# ---------- student view --------------------------------------------------------

@my_bp.route("/my")
@login_required
def my_activity():
    if current_user.role == "admin":
        return redirect(url_for("hub.hub_home"))
    activities = Activity.query.order_by(Activity.held_on, Activity.id).all()
    scores = {r.activity_id: r for r in
              Score.query.filter_by(user_id=current_user.id).all()}

    bars, timeline = [], []
    for a in activities:
        sc = scores.get(a.id)
        bars.append({"title": a.title, "kind": a.kind,
                     "pct": round(100 * (sc.score if sc else 0) / a.max_score, 1)
                     if a.max_score else 0})
        if a.kind == "ctf" and sc and sc.detail:
            try:
                for it in json.loads(sc.detail):
                    timeline.append({"date": (it.get("date") or "")[:10],
                                     "points": it["points"],
                                     "name": it["name"]})
            except (ValueError, KeyError):
                pass
    timeline.sort(key=lambda x: x["date"])

    rows = []
    for a in activities:
        sc = scores.get(a.id)
        rows.append({
            "title": a.title, "kind": a.kind,
            "held_on": a.held_on.strftime("%Y-%m-%d") if a.held_on else "—",
            "score": fmt_score(sc.score if sc else None, a.max_score),
            "note": (sc.detail or "") if sc and a.kind != "ctf" else "",
        })

    sum_pts = sum(scores[a.id].score for a in activities if a.id in scores)
    max_all = sum(a.max_score for a in activities)
    chal_count = 0
    ctf_act = next((a for a in activities if a.kind == "ctf"), None)
    if ctf_act and ctf_act.id in scores and scores[ctf_act.id].detail:
        try:
            chal_count = len(json.loads(scores[ctf_act.id].detail))
        except ValueError:
            chal_count = 0

    return render_template(
        "my_activity.html", rows=rows, bars=bars, timeline=timeline,
        overall_pct=round(100 * sum_pts / max_all, 1) if max_all else 0,
        total_text=f"{fmt_int(sum_pts)} از {fmt_int(max_all)} امتیاز",
        solved=len(timeline), chal_count=chal_count,
        bars_json=json.dumps(bars, ensure_ascii=False),
        timeline_json=json.dumps(timeline, ensure_ascii=False),
    )


# ---------- hub home ------------------------------------------------------------

@hub_bp.route("")
@hub_required
def hub_home():
    students = scoped_students()
    teachers = {t.id: t for t in
                User.query.filter_by(role="admin").order_by(User.username).all()}
    activities = Activity.query.order_by(Activity.held_on, Activity.id).all()
    scores = activity_rows(students, activities)

    act_info = []
    for a in activities:
        vals = [scores[(a.id, s.id)].score for s in students
                if (a.id, s.id) in scores]
        act_info.append({
            "id": a.id, "kind": a.kind, "title": a.title,
            "held_on": a.held_on.strftime("%Y-%m-%d") if a.held_on else "—",
            "max": fmt_int(a.max_score),
            "count": len(vals),
            "avg": round(sum(vals) / len(vals), 1) if vals else 0,
            "avg_pct": round(100 * sum(vals) / len(vals) / a.max_score, 0)
            if vals and a.max_score else 0,
        })

    # per-student totals for the table + chart
    ctf_act = next((a for a in activities if a.kind == "ctf"), None)
    totals = []
    for s in students:
        sum_pts = sum(scores[(a.id, s.id)].score for a in activities
                      if (a.id, s.id) in scores)
        max_all = sum(a.max_score for a in activities)
        ctf_score = None
        if ctf_act and (ctf_act.id, s.id) in scores:
            ctf_score = scores[(ctf_act.id, s.id)].score
        totals.append({
            "s": s,
            "teacher": teacher_name(s.teacher_id, teachers),
            "sum": fmt_int(sum_pts),
            "pct": round(100 * sum_pts / max_all, 1) if max_all else 0,
            "ctf": fmt_score(ctf_score, ctf_act.max_score if ctf_act else 1),
        })
    totals.sort(key=lambda t: -t["pct"])
    return render_template("hub_home.html", activities=act_info, totals=totals,
                           is_super=current_user.is_super,
                           my_students=len(students),
                           acts_json=json.dumps(act_info, ensure_ascii=False))


# ---------- quizzes --------------------------------------------------------------

@hub_bp.route("/quiz/new", methods=["POST"])
@hub_required
def quiz_new():
    title = request.form.get("title", "").strip()
    if not title:
        flash("عنوان کوییز لازم است.", "error")
        return redirect(url_for("hub.hub_home"))
    held = request.form.get("held_on", "").strip()
    max_s = request.form.get("max_score", "20").strip() or "20"
    a = Activity(kind="quiz", title=title,
                 held_on=datetime.strptime(held, "%Y-%m-%d").date()
                 if held else date.today(),
                 max_score=float(parse_score_cell(max_s) or 20),
                 created_by=current_user.id)
    db.session.add(a)
    db.session.commit()
    flash(f"کوییز «{title}» ساخته شد؛ حالا نمره‌ها را وارد کنید.", "ok")
    return redirect(url_for("hub.quiz_view", quiz_id=a.id))


@hub_bp.route("/quiz/<int:quiz_id>", methods=["GET", "POST"])
@hub_required
def quiz_view(quiz_id):
    a = Activity.query.get_or_404(quiz_id)
    if a.kind != "quiz":
        abort(404)
    students = scoped_students()
    if request.method == "POST":
        n_saved, n_blank = 0, 0
        for st in students:
            raw = request.form.get(f"score_{st.id}", "").strip()
            note = request.form.get(f"note_{st.id}", "").strip()[:500]
            val = parse_score_cell(raw)
            row = Score.query.filter_by(activity_id=a.id, user_id=st.id).first()
            if val is None and not note:
                n_blank += 1
                if row:  # explicit empty = keep old (edit is intentional)
                    pass
                continue
            val = max(0.0, min(val if val is not None else 0.0, a.max_score))
            if not row:
                row = Score(activity_id=a.id, user_id=st.id)
                db.session.add(row)
            row.score = val
            row.detail = note
            row.updated_by = current_user.id
            n_saved += 1
        db.session.commit()
        flash(f"{n_saved} نمره ذخیره شد ({n_blank} خالی ماند).", "ok")
        return redirect(url_for("hub.quiz_view", quiz_id=a.id))

    scores = {r.user_id: r for r in
              Score.query.filter_by(activity_id=a.id).all()}
    return render_template("hub_quiz.html", a=a, students=students,
                           scores=scores, fmt_score=fmt_score)


@hub_bp.route("/quiz/<int:quiz_id>/template")
@hub_required
def quiz_template(quiz_id):
    a = Activity.query.get_or_404(quiz_id)
    students = scoped_students()
    wb = Workbook()
    ws = wb.active
    ws.title = "نمره‌دهی"
    _sheet_header(ws, ["شماره دانش‌آموزی", "نام کاربری", "نام و نام خانوادگی",
                       f"نمره (از {fmt_int(a.max_score)})", "توضیح (اختیاری)"])
    scores = {r.user_id: r for r in Score.query.filter_by(activity_id=a.id).all()}
    for st in students:
        sc = scores.get(st.id)
        ws.append([st.student_no or "", st.username, st.full_name,
                   sc.score if sc else None, (sc.detail or "") if sc else ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"quiz-{a.id}-template.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-"
                              "officedocument.spreadsheetml.sheet")


@hub_bp.route("/quiz/<int:quiz_id>/upload", methods=["POST"])
@hub_required
def quiz_upload(quiz_id):
    a = Activity.query.get_or_404(quiz_id)
    f = request.files.get("xlsx")
    if not f or not f.filename.lower().endswith(".xlsx"):
        flash("فایل xlsx انتخاب کنید.", "error")
        return redirect(url_for("hub.quiz_view", quiz_id=a.id))
    by_uname = {s.username: s for s in scoped_students()}
    by_no = {s.student_no: s for s in by_uname.values() if s.student_no}
    wb = load_workbook(f, data_only=True)
    ws = wb.active
    saved = unknown = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [("" if v is None else str(v).strip()) for v in
                (row[:5] if row else [])]
        if len(vals) < 4 or not any(vals):
            continue
        student = by_uname.get(vals[1].lower()) or by_no.get(vals[0])
        sc = parse_score_cell(vals[3])
        if student is None or sc is None:
            unknown += 1
            continue
        sc = max(0.0, min(sc, a.max_score))
        srow = Score.query.filter_by(activity_id=a.id, user_id=student.id).first()
        if not srow:
            srow = Score(activity_id=a.id, user_id=student.id)
            db.session.add(srow)
        srow.score = sc
        srow.detail = vals[4] if len(vals) > 4 else ""
        srow.updated_by = current_user.id
        saved += 1
    db.session.commit()
    flash(f"{saved} نمره از فایل خوانده شد"
          + (f" ({unknown} ردیف ناشناخته/نامعتبر نادیده گرفته شد)." if unknown else "."),
          "ok" if saved else "error")
    return redirect(url_for("hub.quiz_view", quiz_id=a.id))


@hub_bp.route("/quiz/<int:quiz_id>/delete", methods=["POST"])
@hub_required
def quiz_delete(quiz_id):
    a = Activity.query.get_or_404(quiz_id)
    if a.kind == "ctf":
        flash("فعالیت CTF با دکمهٔ همگام‌سازی مدیریت می‌شود.", "error")
        return redirect(url_for("hub.hub_home"))
    Score.query.filter_by(activity_id=a.id).delete()
    db.session.delete(a)
    db.session.commit()
    flash("کوییز حذف شد.", "ok")
    return redirect(url_for("hub.hub_home"))


# ---------- export / sync --------------------------------------------------------

@hub_bp.route("/export")
@hub_required
def export():
    students = scoped_students()
    teachers = {t.id: t for t in User.query.filter_by(role="admin").all()}
    activities = Activity.query.order_by(Activity.held_on, Activity.id).all()
    scores = activity_rows(students, activities)
    wb = build_workbook(students, activities, scores, teachers)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    scope = "all" if current_user.is_super else f"teacher-{current_user.id}"
    return send_file(buf, as_attachment=True,
                     download_name=f"activity-report-{scope}.xlsx",
                     mimetype="application/vnd.openxmlformats-"
                              "officedocument.spreadsheetml.sheet")


@hub_bp.route("/ctf/sync", methods=["POST"])
@hub_required
def ctf_sync():
    return sync_ctf_activity()


# ---------- teachers (super only) ------------------------------------------------

@hub_bp.route("/teachers", methods=["GET", "POST"])
@super_required
def teachers_page():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            uname = request.form.get("username", "").strip().lower()
            fname = request.form.get("full_name", "").strip()
            pw = request.form.get("password", "")
            if not uname.isidentifier() or len(pw) < 8:
                flash("نام کاربری معتبر و رمز حداقل ۸ نویسه لازم است.", "error")
                return redirect(url_for("hub.teachers_page"))
            if User.query.filter_by(username=uname).first():
                flash("این نام کاربری قبلاً ثبت شده است.", "error")
                return redirect(url_for("hub.teachers_page"))
            t = User(role="admin", username=uname, full_name=fname,
                     status=STATUS_APPROVED)
            t.set_password(pw)
            db.session.add(t)
            db.session.commit()
            flash(f"معلم «{uname}» ساخته شد.", "ok")

        elif action == "assign":
            n = 0
            for st in User.query.filter(User.role == "student",
                                        User.status == STATUS_APPROVED):
                new_no = request.form.get(f"no_{st.id}", "").strip()[:32]
                new_t = request.form.get(f"teacher_{st.id}", "").strip()
                tid = int(new_t) if new_t.isdigit() else None
                if new_no != (st.student_no or "") or tid != st.teacher_id:
                    st.student_no, st.teacher_id = new_no, tid
                    n += 1
            db.session.commit()
            flash(f"{n} تغییر ذخیره شد.", "ok")

        elif action == "bulk":
            tmap = {t.username: t.id for t in
                    User.query.filter_by(role="admin")}
            n = bad = 0
            for line in request.form.get("bulk", "").splitlines():
                parts = [p.strip().lower() for p in line.split(",") if p.strip()]
                if len(parts) != 2:
                    if line.strip():
                        bad += 1
                    continue
                st = User.query.filter_by(username=parts[0], role="student").first()
                tid = tmap.get(parts[1])
                if not st or not tid:
                    bad += 1
                    continue
                st.teacher_id = tid
                n += 1
            db.session.commit()
            flash(f"{n} دانشجو تخصیص یافت" + (f" ({bad} خط نامعتبر)." if bad else "."),
                  "ok" if n else "error")
        return redirect(url_for("hub.teachers_page"))

    teachers = User.query.filter_by(role="admin").order_by(User.username).all()
    students = User.query.filter(User.role == "student",
                                 User.status == STATUS_APPROVED)\
        .order_by(User.username).all()
    counts = {t.id: sum(1 for s in students if s.teacher_id == t.id)
              for t in teachers}
    return render_template("hub_teachers.html", teachers=teachers,
                           students=students, counts=counts)
